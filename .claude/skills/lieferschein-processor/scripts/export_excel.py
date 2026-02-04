#!/usr/bin/env python3
"""
Excel Export Script for Lieferschein Data

Exports extracted delivery note data to formatted Excel files.
Supports batch processing and template-based formatting.

Usage:
    python export_excel.py data.json --output result.xlsx
    python export_excel.py data_dir/ --output batch_results.xlsx --batch
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelExporter:
    """
    Export Lieferschein data to formatted Excel files.

    Supports single document export and batch processing with
    professional formatting and conditional styling.
    """

    # Column configuration
    COLUMNS = [
        ('document_number', 'Dokument-Nr.', 15),
        ('delivery_date', 'Lieferdatum', 12),
        ('customer_name', 'Kunde', 25),
        ('customer_address', 'Lieferadresse', 35),
        ('product_type', 'Produkt', 12),
        ('quantity_liters', 'Menge (L)', 12),
        ('tank_level_before', 'Tankstand Vorher', 15),
        ('tank_level_after', 'Tankstand Nachher', 15),
        ('driver_name', 'Fahrer', 18),
        ('customer_signature', 'Unterschrift', 12),
        ('handwritten_notes', 'Bemerkungen', 30),
        ('confidence_score', 'Konfidenz', 10),
        ('needs_review', 'Prüfung erforderlich', 18),
        ('validation_errors', 'Validierungsfehler', 40),
    ]

    # Styles
    HEADER_FILL = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
    REVIEW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    ERROR_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    OK_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self):
        pass

    def export_single(self, data: dict, output_path: str) -> None:
        """
        Export a single Lieferschein record to Excel.

        Args:
            data: Dictionary with extracted data
            output_path: Path for output Excel file
        """
        self.export_batch([data], output_path)

    def export_batch(self, records: list[dict], output_path: str) -> None:
        """
        Export multiple Lieferschein records to Excel.

        Args:
            records: List of dictionaries with extracted data
            output_path: Path for output Excel file
        """
        # Prepare data
        prepared_records = [self._prepare_record(r) for r in records]

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Lieferscheine"

        # Write headers
        for col_idx, (_, header, width) in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.THIN_BORDER
            ws.column_dimensions[cell.column_letter].width = width

        # Write data rows
        for row_idx, record in enumerate(prepared_records, start=2):
            needs_review = record.get('needs_review', False)
            has_errors = bool(record.get('validation_errors'))

            for col_idx, (field, _, _) in enumerate(self.COLUMNS, start=1):
                value = record.get(field, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(vertical='center', wrap_text=True)

                # Apply conditional formatting
                if has_errors:
                    cell.fill = self.ERROR_FILL
                elif needs_review:
                    cell.fill = self.REVIEW_FILL
                elif field == 'needs_review' and not needs_review:
                    cell.fill = self.OK_FILL

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add auto-filter
        ws.auto_filter.ref = ws.dimensions

        # Add summary sheet
        self._add_summary_sheet(wb, prepared_records)

        # Save
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))

        print(f"Exported {len(records)} record(s) to {output_path}")

    def _prepare_record(self, record: dict) -> dict:
        """Prepare a record for Excel export."""
        prepared = record.copy()

        # Convert boolean fields
        if 'customer_signature' in prepared:
            prepared['customer_signature'] = 'Ja' if prepared['customer_signature'] else 'Nein'

        if 'needs_review' in prepared:
            original_value = prepared['needs_review']
            prepared['needs_review'] = 'Ja' if prepared['needs_review'] else 'Nein'
            prepared['_needs_review_bool'] = original_value

        # Convert list fields
        if 'validation_errors' in prepared:
            if isinstance(prepared['validation_errors'], list):
                prepared['validation_errors'] = '; '.join(prepared['validation_errors'])

        # Format confidence score
        if 'confidence_score' in prepared:
            conf = prepared['confidence_score']
            if isinstance(conf, (int, float)):
                prepared['confidence_score'] = f"{conf:.1%}"

        return prepared

    def _add_summary_sheet(self, wb: Workbook, records: list[dict]) -> None:
        """Add a summary statistics sheet."""
        ws = wb.create_sheet(title="Zusammenfassung")

        # Calculate statistics
        total = len(records)
        needs_review = sum(1 for r in records if r.get('_needs_review_bool', r.get('needs_review') == 'Ja'))
        has_errors = sum(1 for r in records if r.get('validation_errors'))
        auto_approved = total - needs_review

        total_liters = sum(
            r.get('quantity_liters', 0) or 0
            for r in records
            if isinstance(r.get('quantity_liters'), (int, float))
        )

        # Write summary
        summary_data = [
            ("Zusammenfassung - Lieferschein-Verarbeitung", ""),
            ("", ""),
            ("Generiert am", datetime.now().strftime("%d.%m.%Y %H:%M")),
            ("", ""),
            ("Statistiken", ""),
            ("Dokumente gesamt", total),
            ("Auto-freigegeben", auto_approved),
            ("Prüfung erforderlich", needs_review),
            ("Mit Fehlern", has_errors),
            ("", ""),
            ("Gesamtmenge (Liter)", f"{total_liters:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")),
        ]

        for row_idx, (label, value) in enumerate(summary_data, start=1):
            cell_label = ws.cell(row=row_idx, column=1, value=label)
            cell_value = ws.cell(row=row_idx, column=2, value=value)

            if row_idx == 1:
                cell_label.font = Font(bold=True, size=14)
            elif label in ["Statistiken", ""]:
                cell_label.font = Font(bold=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def create_template(self, output_path: str) -> None:
        """
        Create an empty Excel template with proper formatting.

        Args:
            output_path: Path for template file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Lieferscheine"

        # Write headers
        for col_idx, (_, header, width) in enumerate(self.COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.THIN_BORDER
            ws.column_dimensions[cell.column_letter].width = width

        # Freeze header row
        ws.freeze_panes = 'A2'

        # Add data validation notes
        ws.cell(row=2, column=1, value="<Beispiel-Datensatz>")
        ws.cell(row=2, column=2, value="2024-01-15")
        ws.cell(row=2, column=3, value="Max Mustermann")
        ws.cell(row=2, column=4, value="Musterstraße 1, 12345 Musterstadt")
        ws.cell(row=2, column=5, value="Heizöl")
        ws.cell(row=2, column=6, value=1500)
        ws.cell(row=2, column=9, value="Hans Schmidt")
        ws.cell(row=2, column=10, value="Ja")
        ws.cell(row=2, column=12, value="95.0%")
        ws.cell(row=2, column=13, value="Nein")

        # Apply example row formatting
        for col_idx in range(1, len(self.COLUMNS) + 1):
            cell = ws.cell(row=2, column=col_idx)
            cell.border = self.THIN_BORDER
            cell.fill = self.OK_FILL

        # Save
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))

        print(f"Template created: {output_path}")


def load_json_data(path: str) -> list[dict]:
    """Load data from JSON file or directory."""
    path = Path(path)

    if path.is_file():
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return data if isinstance(data, list) else [data]

    elif path.is_dir():
        records = []
        for json_file in path.glob('*.json'):
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, list):
                    records.extend(data)
                else:
                    records.append(data)
        return records

    else:
        raise FileNotFoundError(f"Path not found: {path}")


def main():
    """CLI entry point."""
    parser = argparse.ArgumentParser(
        description='Export Lieferschein data to Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s data.json --output result.xlsx
  %(prog)s ./data_dir/ --output batch.xlsx
  %(prog)s --create-template template.xlsx
        """
    )

    parser.add_argument(
        'input',
        nargs='?',
        help='JSON file or directory with data'
    )

    parser.add_argument(
        '--output', '-o',
        required=False,
        help='Output Excel file path'
    )

    parser.add_argument(
        '--create-template',
        metavar='PATH',
        help='Create an empty template file'
    )

    args = parser.parse_args()

    exporter = ExcelExporter()

    # Create template mode
    if args.create_template:
        exporter.create_template(args.create_template)
        return

    # Export mode
    if not args.input:
        parser.error("Input file/directory required unless --create-template is used")

    if not args.output:
        parser.error("--output is required for export")

    try:
        records = load_json_data(args.input)

        if not records:
            print("Warning: No records found to export", file=sys.stderr)
            sys.exit(1)

        exporter.export_batch(records, args.output)

    except FileNotFoundError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"Error parsing JSON: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
